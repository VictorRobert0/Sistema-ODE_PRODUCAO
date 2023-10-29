import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pandas as pd
import pathlib
#import win32com.client as win32

from openpyxl import Workbook, workbook
from openpyxl.cell import cell
from unicodedata import name

#aparencia padrão do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()
        
        

    def layout_config(self):
        self.title("Sistema de cadastro de clientes - Grupo Ode")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self,text="Tema", bg_color="transparent", text_color=['#000', "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self,values=["Light", "Dark"], command=self.change_apm).place(x=50, y=460)
        
        
    def todo_sistema(self, genero_combobox=None):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal",fg_color="teal")
        frame.place(x=0, y=10)
        
        title = ctk.CTkLabel(frame, text="GRUPO ODE - PEÇAS E SERVIÇOS", font=("Century Gothic bold",24), text_color="#fff").place(x=150, y=10)
        
        spam = ctk.CTkLabel(self, text= "Por favor, não se esqueça de preencher os campos abaixo!",text_color=["#000", "#fff"]).place(x=50, y=70)
         
        ficheiro = pathlib.Path("chatoscar.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha=ficheiro.active
            folha['A1']="nome completo"
            folha['B1']="Contato"
            folha['C1']="Data entrada"
            folha['D1']="Genero"
            folha['E1']="Endereço"
            folha['F1']="Observações"
            folha['F1']="Placa"
            folha['F1']="KM Veiculo"
            folha['F1']="Modelo"
            folha['F1']="Versão"
            
            

            ficheiro.save("chatoscar.xlsx")
        
        
        
        
        def submit():
            #Pegando os dados dos entrys
            name = name_value.get()
            contato = contact_value.get()
            dataentrada = dataentrada_value.get()
            gender = genero_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0, END)
            placa = placa_value.get()
            km = km_entry.get(0.0,END)
            fabricacao = fabricacao_value.get()
            versao = versao_value.get()
            
           
            
            
            
            ficheiro = openpyxl.load_workbook('chatoscar.xlsx')
    
            teste = ficheiro.active
            teste.cell(column=1, row=teste.max_row+1, value=name)
            teste.cell(column=2, row=teste.max_row, value=contato)
            teste.cell(column=3, row=teste.max_row, value=dataentrada)
            teste.cell(column=4, row=teste.max_row, value=gender)
            teste.cell(column=5, row=teste.max_row, value=address)
            teste.cell(column=6, row=teste.max_row, value=obs)
            teste.cell(column=7, row=teste.max_row, value=placa)
            teste.cell(column=8, row=teste.max_row, value=km)
            teste.cell(column=9, row=teste.max_row, value=fabricacao)
            teste.cell(column=10, row=teste.max_row, value=versao)
        
            ficheiro.save(r"chatoscar.xlsx")


        # ficheiro.save(r"chatoscar.xlsx")

            messagebox.showinfo("Sistema Grupo ODE", "Dados salvos com sucesso!")
        
            
            


        def clear():
             name_value.set("")
             contact_value.set("")
             dataentrada_value.set("")
             fabricacao_value.set("")
             versao_value.set("")
             address_value.set("")
             obs_entry.delete(0.0, END)
             placa_value.set("")
             km_entry.delete(0.0, END)
            

    
            
  
        
        
        #text variables
        name_value = StringVar()
        contact_value = StringVar()
        dataentrada_value = StringVar()
        address_value = StringVar()
        placa_value = StringVar()
        fabricacao_value = StringVar()
        versao_value = StringVar()

        
        
        
        #Entrys
        name_entry = ctk.CTkEntry(self,width=350, textvariable=name_value,font=("Centure Gothic bold", 16), fg_color="transparent")
        contato_entry = ctk.CTkEntry(self,width=200, textvariable=contact_value, font=("Centure Gothic bold", 16), fg_color="transparent")
        dataentrada_entry = ctk.CTkEntry(self,width=150, textvariable=dataentrada_value, font=("Centure Gothic bold", 16), fg_color="transparent")
        addres_entry = ctk.CTkEntry(self,width=200, textvariable=address_value, font=("Centure Gothic bold", 16), fg_color="transparent")
        placa_entry = ctk.CTkEntry(self,width=100, textvariable=placa_value,font=("Centure Gothic bold", 16), fg_color="transparent")
        fabricacao_entry = ctk.CTkEntry(self,width=150, textvariable=fabricacao_value, font=("Centure Gothic bold", 16), fg_color="transparent")
        versao_entry = ctk.CTkEntry(self,width=150, textvariable=versao_value, font=("Centure Gothic bold", 16), fg_color="transparent")
        
        #Combobox
        genero_combobox =ctk.CTkComboBox(self, values =["Carro", "Moto" ], width=150)
        genero_combobox.set("Carro")
        
        #Entrada de observações
        obs_entry =ctk.CTkTextbox(self,width=400,height=100, border_width=2, fg_color="transparent")
        km_entry =ctk.CTkTextbox(self,width=70,height=0, border_width=2, fg_color="transparent")
   
        
        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome completo",text_color=["#000" , "#fff"])
        lb_contact =ctk.CTkLabel(self, text="Celular",text_color=["#000","#fff"])
        lb_dataentrada =ctk.CTkLabel(self, text="Data de entrada",text_color=["#000","#fff"])
        lb_gender =ctk.CTkLabel(self, text="Veiculo",text_color=["#000","#fff"])
        lb_addres =ctk.CTkLabel(self, text="Endereço",text_color=["#000","#fff"])
        lb_obs =ctk.CTkLabel(self, text="Serviço realizado",text_color=["#000","#fff"])
        lb_placa= ctk.CTkLabel(self, text="Placa", text_color=["#000","#fff"])
        lb_km =ctk.CTkLabel(self, text="KM do veiculo",text_color=["#000","#fff"])
        lb_fabricacao =ctk.CTkLabel(self, text="Fabricação",text_color=["#000","#fff"])
        lb_versao= ctk.CTkLabel(self, text="", text_color=["#000", "#fff"])
        btn_submit = ctk.CTkButton(self,text="SALVAR".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_clear = ctk.CTkButton(self,text="LIMPAR DADOS".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)
        
         #Posicionando elementos na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50,y=150)
        
        lb_contact.place(x=450, y=120)
        contato_entry.place(x=450, y=150)
        
        lb_dataentrada.place(x=300, y=190)
        dataentrada_entry.place(x=300, y=220)
        
        lb_fabricacao.place(x=300, y=260)
        fabricacao_entry.place(x=370, y=260)
        lb_versao.place(x=300, y=260)
        versao_entry.place(x=530, y=260)
        
        placa_entry.place(x=450, y=90)
        lb_placa.place(x=450, y=60)
        
        lb_gender.place(x=500, y=190)
        genero_combobox.place(x=500, y=220)
        
        lb_addres.place(x=50, y=190)
        addres_entry.place(x=50, y=220)
        
        lb_obs.place(x=50, y=300)
        obs_entry.place(x=150, y=300)
        
        lb_km.place(x=555, y=60)
        km_entry.place(x=555, y=90)
        
        
        
        
        
        
        
        
        
        
    
 
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)
        
        



if __name__=="__main__":
    app = App()
    app.mainloop()

